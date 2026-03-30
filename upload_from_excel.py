#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
从Excel读取商品信息并上架到微信小店
"""
import os
import sys
sys.path.insert(0, '/home/wangziyi/.openclaw/workspace/ecommerce_agent')

import openpyxl
from wechat_uploader import SmartWechatUploader


def read_excel_products(excel_path):
    """从Excel读取商品信息"""
    wb = openpyxl.load_workbook(excel_path)
    ws = wb.active
    
    products = []
    
    # 跳过表头，从第2行开始
    for row in range(3, ws.max_row + 1):
        row_data = [cell.value for cell in ws[row]]
        
        # 提取商品信息
        # 列: 0=序号, 1=店铺名, 2=分类, 3=?, 4=?, 5=SKU, 6=品牌, 7=商品名称, 8=副标题, 9=价格
        sku = row_data[5]
        brand = row_data[6]
        name = row_data[7]
        sub_title = row_data[8]
        price = row_data[9]
        
        if name and price:  # 确保有商品名称和价格
            products.append({
                'sku': str(sku) if sku else '',
                'brand': str(brand) if brand else '',
                'name': str(name),
                'sub_title': str(sub_title) if sub_title else '',
                'price': float(price) if price else 0
            })
    
    return products


def upload_products_from_excel(excel_path, img_folder, limit=None):
    """
    从Excel读取商品并上传
    
    Args:
        excel_path: Excel文件路径
        img_folder: 图片文件夹路径
        limit: 上传数量限制（默认全部）
    """
    print("=" * 60)
    print("从Excel上架商品到微信小店")
    print("=" * 60)
    
    # 读取Excel
    print(f"\n读取Excel: {excel_path}")
    products = read_excel_products(excel_path)
    print(f"共读取到 {len(products)} 个商品")
    
    if limit:
        products = products[:limit]
        print(f"限制上传前 {limit} 个")
    
    # 初始化上传器
    uploader = SmartWechatUploader()
    
    # 获取模板图片（因为直接上传图片失败）
    print("\n获取模板图片...")
    try:
        config = uploader.get_template_config()
        main_imgs = config.get('head_imgs', [])[:4]
        detail_imgs = config.get('desc_info_imgs', [])[:10]
        print(f"  模板主图: {len(main_imgs)} 张")
        print(f"  模板详情图: {len(detail_imgs)} 张")
    except Exception as e:
        print(f"  获取模板失败: {e}")
        return
    
    # 上传每个商品
    print("\n" + "=" * 60)
    print("开始上传商品...")
    print("=" * 60)
    
    success_count = 0
    fail_count = 0
    
    for i, product in enumerate(products, 1):
        print(f"\n[{i}/{len(products)}] 上传: {product['name'][:30]}...")
        
        product_info = {
            'name': product['name'],
            'price': product['price'],
            'stock': 100,
            'main_images': main_imgs,
            'detail_images': detail_imgs
        }
        
        try:
            result = uploader.smart_create_and_list(product_info)
            
            if result.get('success'):
                print(f"  ✅ 成功! 商品ID: {result.get('product_id')}")
                success_count += 1
            else:
                print(f"  ❌ 失败: {result.get('message', '未知错误')}")
                fail_count += 1
        except Exception as e:
            print(f"  ❌ 异常: {e}")
            fail_count += 1
        
        # 避免请求过快
        import time
        time.sleep(1)
    
    print("\n" + "=" * 60)
    print("上传完成!")
    print(f"  成功: {success_count}")
    print(f"  失败: {fail_count}")
    print("=" * 60)


def main():
    """主函数"""
    excel_path = '/home/wangziyi/桌面/上架测试货盘.xlsx'
    img_folder = '/home/wangziyi/桌面/上架商品图'
    
    # 可以选择上传数量，默认全部
    # limit = 3  # 只上传前3个
    limit = None  # 全部
    
    upload_products_from_excel(excel_path, img_folder, limit)


if __name__ == '__main__':
    main()
